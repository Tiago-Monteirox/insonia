import sys
sys.path.append('/home/tiago/work/insonia')

import os
import django
from pathlib import Path
from openpyxl import load_workbook  # Utilizando openpyxl para trabalhar com arquivos Excel (xlsx)
from django.core.exceptions import ObjectDoesNotExist
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "insonia.settings")
django.setup()
from lojapp.models import Categoria
from lojapp.models import Produto

project_directory = Path(__file__).resolve().parent.parent
os.chdir(project_directory)

def importar_dados_excel(caminho_arquivo):
    workbook = load_workbook('/home/tiago/Documents/tabela_exportacao_produtos_projeto.xlsx')
    sheet = workbook.active

    cabecalhos = next(sheet.iter_rows(min_row=1, values_only=True))

    for linha in sheet.iter_rows(min_row=2, values_only=True):  # Começa da segunda linha, assumindo que a primeira contém cabeçalhos
        name, categoria, preco_custo, quantidade = linha

        # Encontre ou crie a Categoria
        categoria, created = Categoria.objects.get_or_create(name=categoria)

            # Criar o Produto no banco de dados
        Produto.objects.create(
            name=name,
            preco_custo=preco_custo,
            quantidade=quantidade,
            categoria=categoria,  # Adicione a categoria aqui
            preco_venda=None,
            preco_venda_promocional=None,
            descricao_curta='',
            slug=None,
        )

if __name__ == "__main__":
    caminho_arquivo_excel = '/home/tiago/Documents/tabela_exportacao_produtos_projeto_1.xlsx'
    importar_dados_excel(caminho_arquivo_excel)

print("Importação de dados concluída com sucesso!")